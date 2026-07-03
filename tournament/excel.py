"""Excel workbook builder and its colour/formula helpers."""
import math
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .constants import RR, RR6_ROUNDS
from .utils import fmt, base_level


_BG={'B':'C6EFCE','BB':'BDD7EE','A':'FCE4D6','Open':'FFF2CC','A+BB':'E8D5F5','A+Open':'FFE8D5'}
_FG={'B':'1E4620','BB':'1F4E79','A':'843C0C','Open':'7F6000','A+BB':'4A235A','A+Open':'6B2800'}
_BK={'B':'A9D18E','BB':'9DC3E6','A':'F4B183','Open':'FFD966'}
_PI={'B':'D5F5E3','BB':'D6EAF8','A':'FDEBD0','Open':'FEF9E7'}
HDR=('1F3864','FFFFFF')

def lc(lvl,phase):
    key=lvl.split('+')[0]
    for k in ('BB','B','Open','A'):
        if k==key or k in lvl:
            if 'DE' in phase or 'CHAMP' in phase: return _BK.get(k,'DDDDDD'),_FG.get(k,'000000')
            if phase=='POOL PLAY': return _BG.get(lvl,_BG.get(k,'F2F2F2')),_FG.get(lvl,_FG.get(k,'000000'))
            if phase=='PLAY-IN':   return _PI.get(k,'EEEEEE'),_FG.get(k,'000000')
            return _BK.get(k,'DDDDDD'),_FG.get(k,'000000')
    return 'F2F2F2','000000'


def h2h_round_idx(rr,t1,t2):
    for i,(a,b,_) in enumerate(rr):
        if {a,b}=={t1,t2}: return i
    return None

def make_rank_formula(k,local,team_rows,rr,pc=8):
    my_row=team_rows[local-1]; parts=[]
    for j in range(1,k+1):
        if j==local: continue
        j_row=team_rows[j-1]; r_idx=h2h_round_idx(rr,local,j)
        if r_idx is not None:
            pc_col=get_column_letter(pc+r_idx*2+1)
            hm=f'{pc_col}{my_row}'; hj=f'{pc_col}{j_row}'
        else: hm=hj='0'
        c1=f'(C{j_row}>C{my_row})'; c2=f'((C{j_row}=C{my_row})*(E{j_row}>E{my_row}))'; c3=f'((C{j_row}=C{my_row})*(E{j_row}=E{my_row})*({hj}>{hm}))'
        parts.append(f'(({c1}+{c2}+{c3})>0)')
    return '=1' if not parts else '=1+SUMPRODUCT(('+'+'.join(parts)+')*1)'

def wl_f(letter,col,row):
    c=f'{col}{row}'; return f'IF(LEN(UPPER({c}))<=2,LEN(UPPER({c}))-LEN(SUBSTITUTE(UPPER({c}),"{letter}","")),0)'

def rng(rows,col): return f'{col}{rows[0]}:{col}{rows[-1]}'


# ── Excel builder ──────────────────────────────────────────────────────────────
def build_excel(pools,de_divs,matches,bstruct,combined_info,level_end,de_end_times,warning,counts,
                title=None,stagger_offsets=None):
    wb=Workbook()
    thin=Side(style='thin',color='BBBBBB'); med=Side(style='medium',color='444444')
    tb=Border(left=thin,right=thin,top=thin,bottom=thin); mb=Border(left=med,right=med,top=med,bottom=med)
    def W(ws,r,c,v='',bold=False,bg=None,fg='000000',align='center',brd=None,sz=9,wrap=False,italic=False,mc=None):
        if mc: ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=mc)
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(bold=bold,color=fg,name='Calibri',size=sz,italic=italic)
        if bg: cell.fill=PatternFill('solid',start_color=bg)
        cell.alignment=Alignment(horizontal=align,vertical='center',wrap_text=wrap,indent=(1 if align=='left' else 0))
        if brd: cell.border=brd
        return cell

    # Sheet 1: Full Schedule
    ws=wb.active; ws.title='Full Schedule'; ws.sheet_view.showGridLines=False
    for i,w in enumerate([14,7,14,34,22,22,22,36],1): ws.column_dimensions[get_column_letter(i)].width=w
    r=1
    ws.merge_cells(f'A{r}:H{r}'); W(ws,r,1,(title or 'VOLLEYBALL TOURNAMENT') + ' - FULL SCHEDULE',bold=True,bg=HDR[0],fg=HDR[1],sz=14,align='left'); ws.row_dimensions[r].height=30; r+=1
    np_=sum(1 for m in matches if m.get('is_pool')); nb_=len(matches)-np_
    ws.merge_cells(f'A{r}:H{r}'); W(ws,r,1,f'  9 Courts  |  7:20 AM - 8:30 PM  |  Pool/DE: 45 min  |  Bracket: 60 min  |  {np_} Pool/DE + {nb_} Bracket = {len(matches)} Total',bg='2F5496',fg=HDR[1],sz=10,align='left'); ws.row_dimensions[r].height=18; r+=1
    if warning:
        ws.merge_cells(f'A{r}:H{r}'); W(ws,r,1,f'  WARNING: {warning}',bold=True,bg='FFF2CC',fg='7F4F00',sz=9,align='left',wrap=True); ws.row_dimensions[r].height=22; r+=1
    if de_divs:
        ws.merge_cells(f'A{r}:H{r}')
        de_info='  |  '.join(f'{lvl}: {len(t)} teams (double elimination, random seeding)' for lvl,t in de_divs.items())
        W(ws,r,1,f'  DOUBLE ELIM: {de_info}',bg='E8D5F5',fg='4A235A',sz=9,bold=True,align='left'); ws.row_dimensions[r].height=18; r+=1
    if level_end:
        ws.merge_cells(f'A{r}:H{r}')
        W(ws,r,1,'  POOL PLAY ENDS: '+'  |  '.join(f'{l} ends {fmt(t)}' for l,t in sorted(level_end.items())),bg='E2EFDA',fg='375623',sz=9,bold=True,align='left'); ws.row_dimensions[r].height=18; r+=1
    if stagger_offsets and any(v>0 for v in stagger_offsets.values()):
        ws.merge_cells(f'A{r}:H{r}')
        stag_str='  STAGGERED STARTS: '+' | '.join(f'{l} starts {fmt(t)}' for l,t in sorted(stagger_offsets.items()) if t>0)
        W(ws,r,1,stag_str,bg='E8F4F8',fg='1A5276',sz=9,bold=True,align='left'); ws.row_dimensions[r].height=18; r+=1
    ws.merge_cells(f'A{r}:H{r}'); W(ws,r,1,'  OPEN ELIGIBILITY: Open teams with 2+ losses vs A in pool play cannot qualify for the Open Championship.',bg='FFF2CC',fg='7F4F00',sz=9,bold=True,align='left',wrap=True); ws.row_dimensions[r].height=20; r+=1
    ws.row_dimensions[r].height=8; r+=1
    for i,h in enumerate(['Time','Court','Division','Match','Team A','Team B','Working','Notes'],1): W(ws,r,i,h,bold=True,bg=HDR[0],fg=HDR[1],sz=10,brd=tb)
    ws.row_dimensions[r].height=20; ws.freeze_panes=f'A{r+1}'; r+=1

    prev_tl=None
    for m in matches:
        tl=m['time_label']
        if tl!=prev_tl:
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
            phase=m.get('phase','')
            if m.get('is_de'): dur_str='45 min (DE)'
            elif m.get('is_pool'): dur_str='45 min'
            else: dur_str='60 min'
            W(ws,r,1,f"  \u25b6  {tl}   ends {m['end_label']}   [{phase} - {dur_str}]",bold=True,bg='EBEBEB',fg='1F3864',sz=10,align='left',brd=mb); ws.row_dimensions[r].height=18; r+=1; prev_tl=tl
        lvl=m.get('level',''); phase=m.get('phase','')
        bg_,fg_=lc(lvl,phase); is_b=not m.get('is_pool',True)
        vals=[m['time_label'],f"Court {m['court']}",lvl,m['label'],m['team_a'],m['team_b'],m['team_work'],m.get('note','')]
        for ci,v in enumerate(vals,1):
            W(ws,r,ci,v,bold=(is_b and ci<=4),bg=bg_,fg=fg_ if ci<=3 else ('666666' if ci==8 else '000000'),brd=tb,sz=9,wrap=(ci==8),italic=(ci==8 and bool(v)))
        ws.row_dimensions[r].height=18 if m.get('note') else 15; r+=1

    # Sheet 2: Court Rotation
    ws2=wb.create_sheet('Court Rotation'); ws2.sheet_view.showGridLines=False
    ws2.column_dimensions['A'].width=11
    for i in range(2,12): ws2.column_dimensions[get_column_letter(i)].width=10
    ws2.merge_cells('A1:J1'); W(ws2,1,1,'POOL PLAY - COURT ROTATION (45-min slots)',bold=True,bg=HDR[0],fg=HDR[1],sz=12,align='left'); ws2.row_dimensions[1].height=26; ws2.row_dimensions[2].height=10
    for i,h in enumerate(['Time']+[f'Court {c}' for c in range(1,10)],1): W(ws2,3,i,h,bold=True,bg=HDR[0],fg=HDR[1],sz=10,brd=tb)
    ws2.row_dimensions[3].height=20; rr2=4
    for t in sorted(set(m['start'] for m in matches if m.get('is_pool') or m.get('is_de'))):
        W(ws2,rr2,1,fmt(t),bold=True,bg='F0F0F0',fg='1F3864',brd=tb,sz=9)
        ct_map={m['court']:m['level'] for m in matches if (m.get('is_pool') or m.get('is_de')) and m['start']==t}
        for ct in range(1,10):
            lvl=ct_map.get(ct)
            if lvl: b2,f2=lc(lvl,'POOL PLAY'); W(ws2,rr2,ct+1,lvl[:6],bg=b2,fg=f2,brd=tb,sz=9)
            else: W(ws2,rr2,ct+1,'-',bg='F8F8F8',fg='AAAAAA',brd=tb,sz=9,italic=True)
        ws2.row_dimensions[rr2].height=18; rr2+=1

    # Sheet 3: Bracket Summary
    ws3=wb.create_sheet('Bracket Summary'); ws3.sheet_view.showGridLines=False
    for i,w in enumerate([16,10,12,12,12,14,12],1): ws3.column_dimensions[get_column_letter(i)].width=w
    ws3.merge_cells('A1:G1'); W(ws3,1,1,'BRACKET SUMMARY',bold=True,bg=HDR[0],fg=HDR[1],sz=12,align='left'); ws3.row_dimensions[1].height=26; ws3.row_dimensions[2].height=10
    for i,h in enumerate(['Division','Pools','Advance/Pool','Total','Play-Ins','Format','Direct Seeds'],1): W(ws3,3,i,h,bold=True,bg=HDR[0],fg=HDR[1],sz=10,brd=tb)
    ws3.row_dimensions[3].height=20; br=4
    if de_divs:
        for lvl,teams in de_divs.items():
            n_grp=math.ceil(len(teams)/8); b3,f3=lc(lvl,'DOUBLE ELIM')
            for ci,v in enumerate([lvl,f'{n_grp} groups','N/A (DE)',len(teams),'N/A',f'Double Elimination ({n_grp} courts)','Random'],1):
                W(ws3,br,ci,v,bg=b3,fg=f3 if ci==1 else '000000',bold=(ci==1),brd=tb,sz=10)
            ws3.row_dimensions[br].height=20; br+=1
    for lvl,bs_ in bstruct.items():
        primary=lvl.split('+')[0]; b3,f3=lc(primary,'POOL PLAY'); adv_pp=bs_.get('advances_per_pool',2)
        for ci,v in enumerate([primary,bs_['n_pools'],adv_pp,bs_['total'],bs_['n_playin'],bs_['format'],bs_['n_direct']],1):
            W(ws3,br,ci,v,bg=b3,fg=f3 if ci==1 else '000000',bold=(ci==1),brd=tb,sz=10)
        ws3.row_dimensions[br].height=20; br+=1
        if base_level(lvl).startswith('A+'):
            sec=lvl.split('+')[1]; b3s,f3s=lc(sec,'POOL PLAY')
            for ci,v in enumerate([sec,'(combined)','top eligible','2','0','Championship Match','2'],1):
                W(ws3,br,ci,v,bg=b3s,fg=f3s if ci==1 else '000000',bold=(ci==1),brd=tb,sz=10,italic=True)
            ws3.row_dimensions[br].height=18; br+=1

    ws3.row_dimensions[br].height=10; br+=1
    ws3.merge_cells(f'A{br}:G{br}')
    W(ws3,br,1,'SEEDING: 1st-place finishers seeded first across all pools, then 2nd-place, then 3rd. NCAA bracket: 1v8,4v5,3v6,2v7. 6-team: play-ins 3v6+4v5, seeds 1+2 get byes. DE: random draw within groups, top 2 per group advance to combined championship.',
      bg='F0F0F0',fg='444444',sz=9,align='left',wrap=True); ws3.row_dimensions[br].height=44

    # Sheet 4: Team Tracker
    ws4=wb.create_sheet('Team Tracker'); ws4.sheet_view.showGridLines=False
    ws4.merge_cells('A1:M1'); W(ws4,1,1,'TEAM TRACKER  -  Enter team names and results (WW/WL/LW/LL). Pool Rank and Bracket Seed update automatically.',bold=True,bg=HDR[0],fg=HDR[1],sz=12,align='left'); ws4.row_dimensions[1].height=28
    ws4.merge_cells('A2:M2'); W(ws4,2,1,'W/L: WW=win both  WL/LW=split  LL=lose both  *=working  |  Rank: 1)Wins 2)PtDiff 3)H2H  |  Seeding: 1st-placers by win% then ptdiff, then 2nd, then 3rd.',bg='E2EFDA',fg='375623',sz=9,align='left',italic=True,wrap=True); ws4.row_dimensions[2].height=22; ws4.row_dimensions[3].height=10
    for col,w in [('A',16),('B',20),('C',5),('D',5),('E',10),('F',11),('G',13)]: ws4.column_dimensions[col].width=w
    PC=8; trow=4
    by_level=defaultdict(list)
    for p in pools: by_level[p['level']].append(p)
    pool_meta=defaultdict(list)

    all_pool_levels=list(dict.fromkeys(p['level'] for p in pools))
    for eff_lvl in all_pool_levels:
        lvl_pools=by_level[eff_lvl]; primary=eff_lvl.split('+')[0]; is_combined=('+' in eff_lvl)
        bg_h,fg_h=lc(primary,'POOL PLAY'); bs_lvl=bstruct.get(eff_lvl,{})
        div_label=f'  {eff_lvl} COMBINED POOL PLAY' if is_combined else f'  {primary} DIVISION'
        ws4.merge_cells(f'A{trow}:M{trow}'); W(ws4,trow,1,div_label,bold=True,bg=bg_h,fg=fg_h,sz=11,align='left'); ws4.row_dimensions[trow].height=22; trow+=1
        if is_combined:
            ws4.merge_cells(f'A{trow}:M{trow}')
            sec=eff_lvl.split('+')[1]
            W(ws4,trow,1,f'  {sec} teams are marked with their division. Open eligibility rules apply.',bg='FFF2CC',fg='7F4F00',sz=9,align='left',italic=True); ws4.row_dimensions[trow].height=18; trow+=1

        for p in lvl_pools:
            nm=p['matches']; pool_sz=p['size']; is6=p['is_six']
            origin=p.get('origin',{})
            for mi in range(nm): ws4.column_dimensions[get_column_letter(PC+mi*2)].width=8; ws4.column_dimensions[get_column_letter(PC+mi*2+1)].width=8
            last_col=PC+nm*2-1
            ws4.merge_cells(f'A{trow}:{get_column_letter(last_col)}{trow}')
            snote='(partial RR, 4 matches each, top 4 advance)' if is6 else f'({pool_sz} teams, {nm} matches)'
            W(ws4,trow,1,f'  Pool {p["pool_id"]}  {snote}',bold=True,bg='D9D9D9',fg='1F3864',sz=10,align='left'); ws4.row_dimensions[trow].height=18; trow+=1
            for ci,h in enumerate(['Team ID','Team Name','W','L','Pt Diff','Pool Rank','Bracket Seed'],1): W(ws4,trow,ci,h,bold=True,bg=bg_h,fg=fg_h,sz=9,brd=tb)
            col_hdrs=[('R{} W/L'.format(i+1),'R{} Pts'.format(i+1)) for i in range(nm)] if is6 else [('M{} W/L'.format(i+1),'M{} Pts'.format(i+1)) for i in range(nm)]
            for mi,(wl_h,pt_h) in enumerate(col_hdrs): c=PC+mi*2; W(ws4,trow,c,wl_h,bold=True,bg=bg_h,fg=fg_h,sz=8,brd=tb); W(ws4,trow,c+1,pt_h,bold=True,bg=bg_h,fg=fg_h,sz=8,brd=tb)
            ws4.row_dimensions[trow].height=18; trow+=1
            wl_cols=[get_column_letter(PC+mi*2) for mi in range(nm)]; pd_cols=[get_column_letter(PC+mi*2+1) for mi in range(nm)]
            team_rows_in_pool=list(range(trow,trow+pool_sz))
            pool_meta[eff_lvl].append({'rows':team_rows_in_pool,'size':pool_sz,'matches':nm,'pool_id':p['pool_id'],'is_six':is6})
            row_bgs=['FFFFFF','F7F7F7']
            for ti,team in enumerate(p['teams']):
                rbg=row_bgs[ti%2]; local=ti+1
                team_origin=origin.get(team,primary)
                # Color team ID differently if from secondary division
                team_bg = rbg
                if is_combined and team_origin != primary:
                    team_bg = _BG.get(team_origin, _BG.get(base_level(team_origin), rbg))
                team_label = f'{team} [{team_origin}]' if is_combined else team
                W(ws4,trow,1,team_label,bg=team_bg,fg='000000',brd=tb,sz=9,align='left')
                W(ws4,trow,2,'',bg='FFFDE7',fg='000000',brd=tb,sz=9,align='left')
                wp=[wl_f('W',cl,trow) for cl in wl_cols]; lp=[wl_f('L',cl,trow) for cl in wl_cols]
                W(ws4,trow,3,'=IFERROR('+'+'.join(wp)+',0)',bg=rbg,fg='1E4620',brd=tb,sz=9)
                W(ws4,trow,4,'=IFERROR('+'+'.join(lp)+',0)',bg=rbg,fg='843C0C',brd=tb,sz=9)
                W(ws4,trow,5,'=IFERROR(SUM('+','.join(f'{c}{trow}' for c in pd_cols)+'),0)',bg=rbg,fg='1F4E79',brd=tb,sz=9)
                rf=f'=RANK(C{trow},C{team_rows_in_pool[0]}:C{team_rows_in_pool[-1]},0)' if is6 else make_rank_formula(pool_sz,local,team_rows_in_pool,RR[pool_sz],pc=PC)
                W(ws4,trow,6,rf,bg='FFF9E6',fg='7F6000',brd=tb,sz=9); W(ws4,trow,7,'',bg='E8F0FE',fg='1F4E79',brd=tb,sz=9)
                if is6:
                    for ri,games in enumerate(RR6_ROUNDS):
                        c=PC+ri*2; opps=[b if a==local else a for a,b in games if local in (a,b)]
                        W(ws4,trow,c,'' if opps else 'REST',bg='E8F5E9' if opps else 'F8F8F8',fg='777777',brd=tb,sz=8)
                        W(ws4,trow,c+1,0 if not opps else '',bg='E8F5E9' if opps else 'F8F8F8',fg='444444',brd=tb,sz=9)
                else:
                    for mi in range(1,nm+1):
                        ai_,bi_,wi_=RR[pool_sz][mi-1]; c=PC+(mi-1)*2
                        if local in (ai_,bi_): mbg,wlv,pdv,ital='E8F5E9','','',False
                        elif local==wi_:       mbg,wlv,pdv,ital='EDEDED','*',0,True
                        else:                  mbg,wlv,pdv,ital='F8F8F8','',0,False
                        W(ws4,trow,c,wlv,bg=mbg,fg='777777',brd=tb,sz=9,italic=ital); W(ws4,trow,c+1,pdv if isinstance(pdv,int) else '',bg=mbg,fg='444444',brd=tb,sz=9)
                ws4.row_dimensions[trow].height=17; trow+=1

        # Seeding table
        trow+=1
        ws4.merge_cells(f'A{trow}:G{trow}'); W(ws4,trow,1,f'  {primary} BRACKET SEEDING  (auto-updates)',bold=True,bg=HDR[0],fg=HDR[1],sz=10,align='left'); ws4.row_dimensions[trow].height=20; trow+=1
        for ci,h in enumerate(['Pool','Pool Rank','Team Name','Win%','Pt Diff','Group Seed','Bracket Seed'],1): W(ws4,trow,ci,h,bold=True,bg='2F5496',fg='FFFFFF',sz=9,brd=tb)
        ws4.row_dimensions[trow].height=18; trow+=1
        adv_pp=bs_lvl.get('advances_per_pool',2); group_trows=defaultdict(list)
        grp_bgs=['DEEAF1','EAF0FF','F0FFF0']
        for pool_rank in range(1,adv_pp+1):
            grp_bg=grp_bgs[min(pool_rank-1,2)]
            for pm in pool_meta[eff_lvl]:
                rows=pm['rows']
                f_rng=rng(rows,'F'); b_rng=rng(rows,'B'); c_rng=rng(rows,'C'); d_rng=rng(rows,'D'); e_rng=rng(rows,'E')
                wf=f'IFERROR(INDEX({c_rng},MATCH({pool_rank},{f_rng},0)),0)'; lf=f'IFERROR(INDEX({d_rng},MATCH({pool_rank},{f_rng},0)),0)'
                ptf=f'IFERROR(INDEX({e_rng},MATCH({pool_rank},{f_rng},0)),0)'; nf=f'=IFERROR(INDEX({b_rng},MATCH({pool_rank},{f_rng},0)),"")'
                pctf=f'=IF(({wf}+{lf})=0,0,{wf}/({wf}+{lf}))'
                W(ws4,trow,1,f'Pool {pm["pool_id"]}',bg=grp_bg,fg='1F3864',brd=tb,sz=9,align='center'); W(ws4,trow,2,pool_rank,bg=grp_bg,fg='1F3864',brd=tb,sz=9,align='center')
                W(ws4,trow,3,nf,bg='FFFDE7',fg='000000',brd=tb,sz=9,align='left'); W(ws4,trow,4,pctf,bg=grp_bg,fg='1F3864',brd=tb,sz=9); ws4.cell(trow,4).number_format='0.0%'
                W(ws4,trow,5,f'={ptf}',bg=grp_bg,fg='1F4E79',brd=tb,sz=9); W(ws4,trow,6,'',bg='FFF9E6',fg='7F6000',brd=tb,sz=9); W(ws4,trow,7,'',bg='E8F0FE',fg='1F4E79',brd=tb,sz=9)
                group_trows[pool_rank].append(trow); ws4.row_dimensions[trow].height=17; trow+=1
        bso=0
        for grp_num in range(1,adv_pp+1):
            g_rows=group_trows.get(grp_num,[])
            if not g_rows: continue
            for r_ in g_rows:
                parts=[]
                for r2 in g_rows:
                    if r2==r_: continue
                    parts.append(f'((D{r2}>D{r_})+((D{r2}=D{r_})*(E{r2}>E{r_}))>0)')
                gsf=('=1+SUMPRODUCT(('+'+'.join(parts)+')*1)') if parts else '=1'
                W(ws4,r_,6,gsf,bg='FFF9E6',fg='7F6000',brd=tb,sz=9); W(ws4,r_,7,f'={bso}+F{r_}',bg='C9DAF8',fg='1F4E79',brd=tb,sz=9,bold=True)
            bso+=len(g_rows)
        ws4.row_dimensions[trow].height=16; trow+=1

    # DE division note in tracker
    if de_divs:
        for lvl,teams in de_divs.items():
            ws4.merge_cells(f'A{trow}:M{trow}')
            W(ws4,trow,1,f'  {lvl} DIVISION: {len(teams)} teams in DOUBLE ELIMINATION format. Seeding by random draw. See Full Schedule tab for bracket.',
              bold=True,bg=lc(lvl,'DE')[0],fg=lc(lvl,'DE')[1],sz=10,align='left')
            ws4.row_dimensions[trow].height=22; trow+=1
        ws4.row_dimensions[trow].height=12; trow+=1

    # Sheet 5: Open Eligibility
    ws5=wb.create_sheet('Open Eligibility'); ws5.sheet_view.showGridLines=False
    for i,w in enumerate([16,14,30,22,14,28],1): ws5.column_dimensions[get_column_letter(i)].width=w
    ws5.merge_cells('A1:F1'); W(ws5,1,1,'OPEN CHAMPIONSHIP - ELIGIBILITY TRACKER',bold=True,bg=HDR[0],fg=HDR[1],sz=12,align='left'); ws5.row_dimensions[1].height=26
    ws5.merge_cells('A2:F2')
    combined_open=any('Open' in k.split('+')[-1] for k in combined_info.keys())
    note_text=('A and Open play in COMBINED POOLS. Open teams in combined pools must be tracked separately. '
               'Open teams with 2+ losses vs A cannot qualify for the Open Championship (may enter A bracket).'
               if combined_open else
               'A and Open always have SEPARATE brackets. Open teams with 2+ losses vs A cannot qualify for the Open Championship.')
    W(ws5,2,1,note_text,bg='FFF2CC',fg='7F4F00',sz=9,bold=True,align='left',wrap=True); ws5.row_dimensions[2].height=32

    # Sheet 6: Format Guide
    ws6=wb.create_sheet('Format Guide'); ws6.sheet_view.showGridLines=False
    ws6.column_dimensions['A'].width=28; ws6.column_dimensions['B'].width=64
    ws6.merge_cells('A1:B1'); W(ws6,1,1,'TOURNAMENT FORMAT GUIDE',bold=True,bg=HDR[0],fg=HDR[1],sz=13,align='left'); ws6.row_dimensions[1].height=28
    thin_b=Border(left=thin,right=thin,top=thin,bottom=thin)
    guide=[('TIMING','',True,HDR[0],HDR[1]),
           ('Start / End','7:20 AM to 8:30 PM (hard cutoff)',False,None,'000000'),
           ('Pool Matches','45 minutes each',False,None,'000000'),
           ('Bracket Matches','60 minutes each',False,None,'000000'),
           ('DE Matches','45 minutes each (same as pool play)',False,None,'000000'),
           ('All Start Together','All pool play and DE brackets start at 7:20 AM simultaneously.',False,'DEEAF1','1F4E79'),
           ('','',False,None,'000000'),
           ('POOL PLAY','',True,HDR[0],HDR[1]),
           ('Divisions','4 SEPARATE brackets always: B, BB, A, Open.',False,None,'000000'),
           ('Combining','If Open or BB has 2-5 teams, combined with A for pool play. A and Open/BB still have separate brackets after.',False,'FFF2CC','7F4F00'),
           ('Pool Sizes','Max 9 pools. Prefer 5→4→3. 6 teams = 1 pool partial RR (top 4 advance).',False,None,'000000'),
           ('Qualify','Top 2/pool (base) or top 3/pool (expanded if time allows).',False,None,'000000'),
           ('','',False,None,'000000'),
           ('DOUBLE ELIMINATION','',True,'4A235A','FFFFFF'),
           ('Trigger','Pool play is always preferred. Double elimination is used only when pool play would require more than 9 simultaneous courts.',False,'E8D5F5','4A235A'),
           ('Format','Teams split into groups of ≤8. Each group runs full DE bracket on 1 dedicated court (all sequential).',False,'E8D5F5','4A235A'),
           ('Match Duration','45 minutes per match.',False,'E8D5F5','4A235A'),
           ('Seeding','Random draw within each group.',False,'E8D5F5','4A235A'),
           ('Grand Final Reset','If LB team wins the Grand Final, a reset match is played.',False,'E8D5F5','4A235A'),
           ('Combined Champ','Top 2 per group advance. 6-team combined championship: play-ins (3v6,4v5) + SF + Final.',False,'E8D5F5','4A235A'),
           ('','',False,None,'000000'),
           ('BRACKET FORMATS','',True,HDR[0],HDR[1]),
           ('4 pools→8-team','QF+SF+Final. NCAA seeding: 1v8, 4v5, 3v6, 2v7.',False,None,'000000'),
           ('3 pools→6-team','Play-ins (3v6,4v5) + SF + Final. Seeds 1,2 get byes.',False,None,'000000'),
           ('2 pools→4-team','SF + Final.',False,None,'000000'),
           ('Expand if time','3 pools→9-team (1 play-in), 4 pools→12-team (4 play-ins), 2 pools→6-team.',False,'E2EFDA','375623'),
           ('','',False,None,'000000'),
           ('OPEN ELIGIBILITY','',True,'C00000','FFFFFF'),
           ('Rule','Open teams with 2+ losses vs A cannot qualify for the Open Championship.',False,'FFF2CC','7F4F00'),
           ('Note','A and Open always have separate brackets even when in combined pool play.',False,'FFF2CC','7F4F00')]
    for ri,(a,bv,bold,bg,fg) in enumerate(guide,2):
        ws6.row_dimensions[ri].height=8 if not a else 20
        if not a: continue
        for ci,v in enumerate([a,bv],1): W(ws6,ri,ci,v,bold=bold,bg=bg,fg=fg,brd=(thin_b if a else None),sz=10,align='left',wrap=True)
    return wb
